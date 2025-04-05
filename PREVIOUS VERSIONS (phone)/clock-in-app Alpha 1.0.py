import os
import tkinter as tk
from tkinter import messagebox, StringVar
from openpyxl import Workbook, load_workbook
import datetime

# Global constants
EXCEL_FILE = 'attendance.xlsx'  # For daily attendance
EMPLOYEE_DATA_FILE = 'employee_data.xlsx'  # For employee data

# Initialize Excel with a new sheet for each day
def init_attendance_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)
    
    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
        wb.save(EXCEL_FILE)

# Initialize employee data Excel
def init_employee_data_excel():
    if not os.path.exists(EMPLOYEE_DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "EmployeeData"
        ws.append(["First Name", "Last Name", "ImagePath", "FaceEncoding"])
        wb.save(EMPLOYEE_DATA_FILE)

# Load employee names for the dropdown list
def load_employee_names():
    employee_names = []
    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name, last_name, *_ = row
        full_name = f"{first_name} {last_name}"
        employee_names.append(full_name)

    return employee_names

# Update employee action (clock-in, break start, etc.)
def update_employee_action(action_type):
    full_name = employee_name_var.get()

    if not full_name:
        messagebox.showerror("Error", "Please select an employee")
        return

    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # Check if today's date already has a sheet; if not, create it
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
    else:
        ws = wb[current_date]

    # Check if the employee already has an entry in today's sheet
    employee_row = None
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == full_name:
            employee_row = row
            break

    # If employee doesn't have an entry, add a new row
    if not employee_row:
        employee_row = [full_name, None, None, None, None]  # Initialize the row with empty shift data
        ws.append(employee_row)
        row_index = ws.max_row  # The index of the newly appended row
    else:
        row_index = employee_row[0].row  # Get the existing row index

    # Update the appropriate column based on the action type
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if action_type == "Clock In":
        ws.cell(row=row_index, column=2).value = now
    elif action_type == "Break Start":
        ws.cell(row=row_index, column=3).value = now
    elif action_type == "Break End":
        ws.cell(row=row_index, column=4).value = now
    elif action_type == "Shift End":
        ws.cell(row=row_index, column=5).value = now

    # Save the workbook
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Success", f"{action_type} recorded for {full_name}")

# Tkinter GUI Setup
app = tk.Tk()
app.title("Employee Attendance System")

# Initialize Excel files
init_employee_data_excel()  # Ensure employee data file is created first
employee_names = load_employee_names()  # Now load employee names
init_attendance_excel()  # Then initialize the attendance file

# Initialize the employee name variable
employee_name_var = StringVar(app)

# Set the initial value of employee_name_var based on the available names
if employee_names:
    employee_name_var.set(employee_names[0])  # Set to the first employee name if available
else:
    employee_name_var.set("")  # Or set it to an empty string if there are no employees

# Create dropdown menu for employee selection
employee_dropdown = tk.OptionMenu(app, employee_name_var, *employee_names)
employee_dropdown.pack(pady=10)

# Main menu buttons
tk.Button(app, text="Clock In", command=lambda: update_employee_action("Clock In")).pack(pady=10)
tk.Button(app, text="Break Start", command=lambda: update_employee_action("Break Start")).pack(pady=10)
tk.Button(app, text="Break End", command=lambda: update_employee_action("Break End")).pack(pady=10)
tk.Button(app, text="Shift End", command=lambda: update_employee_action("Shift End")).pack(pady=10)

# New Employee Registration
tk.Label(app, text="First Name:").pack(pady=5)
first_name_entry = tk.Entry(app)
first_name_entry.pack(pady=5)

tk.Label(app, text="Last Name:").pack(pady=5)
last_name_entry = tk.Entry(app)
last_name_entry.pack(pady=5)

def register_new_employee():
    first_name = first_name_entry.get()
    last_name = last_name_entry.get()

    if not first_name or not last_name:
        messagebox.showerror("Error", "Please enter both first and last names")
        return

    # Save to employee data Excel
    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active
    ws.append([first_name, last_name, "", ""])  # ImagePath and FaceEncoding are left empty
    wb.save(EMPLOYEE_DATA_FILE)

    # Update the dropdown menu to include the new employee
    update_employee_dropdown()

    messagebox.showinfo("Success", f"Employee {first_name} {last_name} registered")
    first_name_entry.delete(0, tk.END)
    last_name_entry.delete(0, tk.END)

def update_employee_dropdown():
    # Reload employee names and update the dropdown
    employee_names = load_employee_names()
    employee_name_var.set(employee_names[0] if employee_names else "")
    employee_dropdown['menu'].delete(0, 'end')  # Clear the current menu options

    for name in employee_names:
        employee_dropdown['menu'].add_command(label=name, command=tk._setit(employee_name_var, name))

tk.Button(app, text="Register New Employee", command=register_new_employee).pack(pady=20)

app.mainloop()
