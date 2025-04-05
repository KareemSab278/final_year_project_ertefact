import os
import tkinter as tk
from tkinter import messagebox, StringVar, Listbox, Scrollbar
from openpyxl import Workbook, load_workbook
import datetime

# Global constants
EXCEL_FILE = 'attendance.xlsx'  # For daily attendance
EMPLOYEE_DATA_FILE = 'employee_data.xlsx'  # For employee data

# Initialize Excel with a new sheet for each day
def init_attendance_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)
   
    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
   
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
        wb.save(EXCEL_FILE)

# Initialize employee data Excel
def init_employee_data_excel():
    if not os.path.exists(EMPLOYEE_DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "EmployeeData"
        ws.append(["First Name", "Last Name"])
        wb.save(EMPLOYEE_DATA_FILE)

# Load employee names for the listbox
def load_employee_names():
    employee_names = []
    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name, last_name, *_ = row
        full_name = f"{first_name} {last_name}"
        employee_names.append(full_name)

    return employee_names

# Update employee action (clock-in, break start, etc.)
def update_employee_action(action_type):
    selected_index = employee_listbox.curselection()
   
    if not selected_index:
        messagebox.showerror("Error", "Please select an employee")
        return

    full_name = employee_listbox.get(selected_index)

    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # Check if today's date already has a sheet; if not, create it
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
    else:
        ws = wb[current_date]

    # Check if the employee already has an entry in today's sheet
    employee_row = None
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == full_name:
            employee_row = row
            break

    # If employee doesn't have an entry, add a new row
    if not employee_row:
        employee_row = [full_name, None, None, None, None]  # Initialize the row with empty shift data
        ws.append(employee_row)
        row_index = ws.max_row  # The index of the newly appended row
    else:
        row_index = employee_row[0].row  # Get the existing row index

    # Update the appropriate column based on the action type
    if action_type == "Clock In":
        ws.cell(row=row_index, column=2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Break Start":
        ws.cell(row=row_index, column=3).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Break End":
        ws.cell(row=row_index, column=4).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Shift End":
        ws.cell(row=row_index, column=5).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Save the workbook
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Success", f"{action_type} recorded for {full_name}")

    # Call the summary functions after each update
    update_daily_summary()
    update_weekly_summary()

def update_daily_summary():
    wb = load_workbook(EXCEL_FILE)
    summary_sheet = wb.create_sheet("Daily Summary") if "Daily Summary" not in wb.sheetnames else wb["Daily Summary"]

    # Clear existing content
    for row in summary_sheet.iter_rows(min_row=1, max_col=3, max_row=summary_sheet.max_row):
        for cell in row:
            cell.value = None

    # Write header
    summary_sheet.append(["Date", "Total Hours Worked (with Breaks)", "Total Hours Worked (without Breaks)"])

    for sheet in wb.sheetnames:
        if sheet != "Daily Summary":
            total_hours_with_breaks = 0
            total_hours_without_breaks = 0
           
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                shift_start = row[1]
                break_start = row[2]
                break_end = row[3]
                shift_end = row[4]

                if shift_start and shift_end:
                    start_time = datetime.datetime.strptime(shift_start, '%Y-%m-%d %H:%M:%S')
                    end_time = datetime.datetime.strptime(shift_end, '%Y-%m-%d %H:%M:%S')
                    worked_time = (end_time - start_time).total_seconds() / 3600

                    total_hours_without_breaks += worked_time

                    if break_start and break_end:
                        break_start_time = datetime.datetime.strptime(break_start, '%Y-%m-%d %H:%M:%S')
                        break_end_time = datetime.datetime.strptime(break_end, '%Y-%m-%d %H:%M:%S')
                        break_time = (break_end_time - break_start_time).total_seconds() / 3600
                        worked_time -= break_time

                    total_hours_with_breaks += worked_time
           
            date = sheet
            summary_sheet.append([date, total_hours_with_breaks, total_hours_without_breaks])

    wb.save(EXCEL_FILE)

def update_weekly_summary():
    wb = load_workbook(EXCEL_FILE)
    summary_sheet = wb.create_sheet("Weekly Summary") if "Weekly Summary" not in wb.sheetnames else wb["Weekly Summary"]

    # Clear existing content
    for row in summary_sheet.iter_rows(min_row=1, max_col=3, max_row=summary_sheet.max_row):
        for cell in row:
            cell.value = None

    # Write header
    summary_sheet.append(["Week", "Total Hours Worked (with Breaks)", "Total Hours Worked (without Breaks)"])
   
    current_week = datetime.datetime.now().isocalendar()[1]
    total_hours_with_breaks = 0
    total_hours_without_breaks = 0

    for sheet in wb.sheetnames:
        if sheet != "Daily Summary" and sheet != "Weekly Summary":
            date_obj = datetime.datetime.strptime(sheet, '%Y-%m-%d')
            week_number = date_obj.isocalendar()[1]

            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                shift_start = row[1]
                break_start = row[2]
                break_end = row[3]
                shift_end = row[4]

                if shift_start and shift_end:
                    start_time = datetime.datetime.strptime(shift_start, '%Y-%m-%d %H:%M:%S')
                    end_time = datetime.datetime.strptime(shift_end, '%Y-%m-%d %H:%M:%S')
                    worked_time = (end_time - start_time).total_seconds() / 3600

                    total_hours_without_breaks += worked_time

                    if break_start and break_end:
                        break_start_time = datetime.datetime.strptime(break_start, '%Y-%m-%d %H:%M:%S')
                        break_end_time = datetime.datetime.strptime(break_end, '%Y-%m-%d %H:%M:%S')
                        break_time = (break_end_time - break_start_time).total_seconds() / 3600
                        worked_time -= break_time

                    total_hours_with_breaks += worked_time

            # Write to weekly summary if the week number matches
            if week_number == current_week:
                summary_sheet.append([current_week, total_hours_with_breaks, total_hours_without_breaks])

    wb.save(EXCEL_FILE)

# Tkinter GUI Setup
app = tk.Tk()
app.title("Attendance System")

# Initialize Excel files
init_employee_data_excel()  # Ensure employee data file is created first
employee_names = load_employee_names()  # Now load employee names
init_attendance_excel()  # Then initialize the attendance file

# Listbox for employee selection
employee_listbox = Listbox(app, height=10, width=50, font=("Helvetica", 14))  # Increase font size
employee_listbox.pack(pady=10)

# Add a scrollbar to the listbox
scrollbar = Scrollbar(app)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
employee_listbox.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=employee_listbox.yview)

# Populate the listbox with employee names
for name in employee_names:
    employee_listbox.insert(tk.END, name)

# Main menu buttons
tk.Button(app, text="Clock In", command=lambda: update_employee_action("Clock In")).pack(pady=10)
tk.Button(app, text="Break Start", command=lambda: update_employee_action("Break Start")).pack(pady=10)
tk.Button(app, text="Break End", command=lambda: update_employee_action("Break End")).pack(pady=10)
tk.Button(app, text="Shift End", command=lambda: update_employee_action("Shift End")).pack(pady=10)

# Manual schedule input button
tk.Button(app, text="Manual Schedule Input", command=lambda: open_schedule_input()).pack(pady=20)

# Function to open manual schedule input window
def open_schedule_input():
    schedule_window = tk.Toplevel(app)
    schedule_window.title("Manual Schedule Input")
    schedule_window.geometry("400x600")  # Adjust the size as per your phone's screen

    # Use a frame to organize Listbox and Scrollbar
    frame_listbox = tk.Frame(schedule_window)
    frame_listbox.pack(pady=10)

    tk.Label(schedule_window, text="Select Employee for Schedule", font=("Helvetica", 14)).pack(pady=10)

    # Listbox for employee selection
    schedule_employee_listbox = Listbox(frame_listbox, height=10, width=40, font=("Helvetica", 14))  # Adjust width
    schedule_employee_listbox.pack(side=tk.LEFT)

    # Add a scrollbar to the listbox inside the frame
    schedule_scrollbar = Scrollbar(frame_listbox)
    schedule_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    schedule_employee_listbox.config(yscrollcommand=schedule_scrollbar.set)
    schedule_scrollbar.config(command=schedule_employee_listbox.yview)

    # Populate the listbox with employee names
    for name in employee_names:
        schedule_employee_listbox.insert(tk.END, name)

    tk.Label(schedule_window, text="Select Your Schedule", font=("Helvetica", 14)).pack(pady=10)

    # Time Input and Break selection organized with grid layout for better alignment
    tk.Label(schedule_window, text="Start Hour").grid(row=1, column=0)
    start_hour_slider = tk.Scale(schedule_window, from_=1, to=12, orient=tk.HORIZONTAL)
    start_hour_slider.grid(row=2, column=0)

    tk.Label(schedule_window, text="Start Minute").grid(row=1, column=1)
    start_minute_slider = tk.Scale(schedule_window, from_=0, to=45, resolution=15, orient=tk.HORIZONTAL)
    start_minute_slider.grid(row=2, column=1)

    # AM/PM selection for start time
    start_period_var = StringVar()
    frame1 = tk.Frame(schedule_window)
    frame1.grid(row=3, column=0, columnspan=2)
    tk.Radiobutton(frame1, text="AM", variable=start_period_var, value="AM").pack(side=tk.LEFT)
    tk.Radiobutton(frame1, text="PM", variable=start_period_var, value="PM").pack(side=tk.LEFT)

    # End Time Input
    tk.Label(schedule_window, text="End Hour").grid(row=4, column=0)
    end_hour_slider = tk.Scale(schedule_window, from_=1, to=12, orient=tk.HORIZONTAL)
    end_hour_slider.grid(row=5, column=0)

    tk.Label(schedule_window, text="End Minute").grid(row=4, column=1)
    end_minute_slider = tk.Scale(schedule_window, from_=0, to=45, resolution=15, orient=tk.HORIZONTAL)
    end_minute_slider.grid(row=5, column=1)

    # AM/PM selection for end time
    end_period_var = StringVar()
    frame2 = tk.Frame(schedule_window)
    frame2.grid(row=6, column=0, columnspan=2)
    tk.Radiobutton(frame2, text="AM", variable=end_period_var, value="AM").pack(side=tk.LEFT)
    tk.Radiobutton(frame2, text="PM", variable=end_period_var, value="PM").pack(side=tk.LEFT)

    # Checkbox for break
    break_var = tk.BooleanVar()
    tk.Checkbutton(schedule_window, text="WILL TAKE 30 MIN BREAK", variable=break_var).pack(pady=10)

    # Submit Button
    tk.Button(schedule_window, text="Submit Schedule", command=lambda: submit_schedule(schedule_employee_listbox)).pack(pady=20)

    # Main Screen button
    tk.Button(schedule_window, text="Main Screen", command=schedule_window.destroy).pack(pady=10)

    schedule_window.transient(app)
    schedule_window.grab_set()
    app.wait_window(schedule_window)

# Function to handle the submission of the manual schedule
def submit_schedule(schedule_employee_listbox):
    selected_index = schedule_employee_listbox.curselection()
    if not selected_index:
        messagebox.showerror("Error", "Please select an employee")
        return

    full_name = schedule_employee_listbox.get(selected_index)

    # Get time inputs
    start_hour = start_hour_slider.get()
    start_minute = start_minute_slider.get()
    start_period = start_period_var.get()
    end_hour = end_hour_slider.get()
    end_minute = end_minute_slider.get()
    end_period = end_period_var.get()
    takes_break = break_var.get()

    # Convert hours and minutes to 24-hour format
    if start_period == "PM" and start_hour < 12:
        start_hour += 12
    elif start_period == "AM" and start_hour == 12:
        start_hour = 0

    if end_period == "PM" and end_hour < 12:
        end_hour += 12
    elif end_period == "AM" and end_hour == 12:
        end_hour = 0

    start_time = datetime.time(start_hour, start_minute)
    end_time = datetime.time(end_hour, end_minute)

    # Here you can save the schedule to a file or database as needed
    # For demonstration, we will just show a message box
    messagebox.showinfo("Schedule Submitted", f"Schedule for {full_name}:\nStart: {start_time}\nEnd: {end_time}\nBreak: {takes_break}")

# Start the Tkinter main loop
app.mainloop()

