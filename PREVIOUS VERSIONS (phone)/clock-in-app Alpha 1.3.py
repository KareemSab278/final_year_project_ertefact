import os
import tkinter as tk
from tkinter import messagebox, StringVar, Listbox, Scrollbar
from openpyxl import Workbook, load_workbook
import datetime

# Global constants
EXCEL_FILE = 'attendance.xlsx'  # For daily attendance
EMPLOYEE_DATA_FILE = 'employee_data.xlsx'  # For employee data

# Initialize Excel with a new sheet for each day
def init_attendance_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)
    
    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
        wb.save(EXCEL_FILE)

# Initialize employee data Excel
def init_employee_data_excel():
    if not os.path.exists(EMPLOYEE_DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "EmployeeData"
        ws.append(["First Name", "Last Name"])
        wb.save(EMPLOYEE_DATA_FILE)

# Load employee names for the listbox
def load_employee_names():
    employee_names = []
    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name, last_name, *_ = row
        full_name = f"{first_name} {last_name}"
        employee_names.append(full_name)

    return employee_names

# Update employee action (clock-in, break start, etc.)
def update_employee_action(action_type):
    selected_index = employee_listbox.curselection()
    
    if not selected_index:
        messagebox.showerror("Error", "Please select an employee")
        return

    full_name = employee_listbox.get(selected_index)

    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # Check if today's date already has a sheet; if not, create it
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
    else:
        ws = wb[current_date]

    # Check if the employee already has an entry in today's sheet
    employee_row = None
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == full_name:
            employee_row = row
            break

    # If employee doesn't have an entry, add a new row
    if not employee_row:
        employee_row = [full_name, None, None, None, None]  # Initialize the row with empty shift data
        ws.append(employee_row)
        row_index = ws.max_row  # The index of the newly appended row
    else:
        row_index = employee_row[0].row  # Get the existing row index

    # Update the appropriate column based on the action type
    if action_type == "Clock In":
        ws.cell(row=row_index, column=2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Break Start":
        ws.cell(row=row_index, column=3).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Break End":
        ws.cell(row=row_index, column=4).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif action_type == "Shift End":
        ws.cell(row=row_index, column=5).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Save the workbook
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Success", f"{action_type} recorded for {full_name}")

    # Call the summary functions after each update
    update_daily_summary()
    update_weekly_summary()

def update_daily_summary():
    wb = load_workbook(EXCEL_FILE)
    summary_sheet = wb.create_sheet("Daily Summary") if "Daily Summary" not in wb.sheetnames else wb["Daily Summary"]

    # Clear existing content
    for row in summary_sheet.iter_rows(min_row=1, max_col=3, max_row=summary_sheet.max_row):
        for cell in row:
            cell.value = None

    # Write header
    summary_sheet.append(["Date", "Total Hours Worked (with Breaks)", "Total Hours Worked (without Breaks)"])

    for sheet in wb.sheetnames:
        if sheet != "Daily Summary":
            total_hours_with_breaks = 0
            total_hours_without_breaks = 0
            
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                shift_start = row[1]
                break_start = row[2]
                break_end = row[3]
                shift_end = row[4]

                if shift_start and shift_end:
                    start_time = datetime.datetime.strptime(shift_start, '%Y-%m-%d %H:%M:%S')
                    end_time = datetime.datetime.strptime(shift_end, '%Y-%m-%d %H:%M:%S')
                    worked_time = (end_time - start_time).total_seconds() / 3600

                    total_hours_without_breaks += worked_time

                    if break_start and break_end:
                        break_start_time = datetime.datetime.strptime(break_start, '%Y-%m-%d %H:%M:%S')
                        break_end_time = datetime.datetime.strptime(break_end, '%Y-%m-%d %H:%M:%S')
                        break_time = (break_end_time - break_start_time).total_seconds() / 3600
                        worked_time -= break_time

                    total_hours_with_breaks += worked_time
            
            date = sheet
            summary_sheet.append([date, total_hours_with_breaks, total_hours_without_breaks])

    wb.save(EXCEL_FILE)

def update_weekly_summary():
    wb = load_workbook(EXCEL_FILE)
    summary_sheet = wb.create_sheet("Weekly Summary") if "Weekly Summary" not in wb.sheetnames else wb["Weekly Summary"]

    # Clear existing content
    for row in summary_sheet.iter_rows(min_row=1, max_col=3, max_row=summary_sheet.max_row):
        for cell in row:
            cell.value = None

    # Write header
    summary_sheet.append(["Week", "Total Hours Worked (with Breaks)", "Total Hours Worked (without Breaks)"])
    
    current_week = datetime.datetime.now().isocalendar()[1]
    total_hours_with_breaks = 0
    total_hours_without_breaks = 0

    for sheet in wb.sheetnames:
        if sheet != "Daily Summary" and sheet != "Weekly Summary":
            date_obj = datetime.datetime.strptime(sheet, '%Y-%m-%d')
            week_number = date_obj.isocalendar()[1]

            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                shift_start = row[1]
                break_start = row[2]
                break_end = row[3]
                shift_end = row[4]

                if shift_start and shift_end:
                    start_time = datetime.datetime.strptime(shift_start, '%Y-%m-%d %H:%M:%S')
                    end_time = datetime.datetime.strptime(shift_end, '%Y-%m-%d %H:%M:%S')
                    worked_time = (end_time - start_time).total_seconds() / 3600

                    total_hours_without_breaks += worked_time

                    if break_start and break_end:
                        break_start_time = datetime.datetime.strptime(break_start, '%Y-%m-%d %H:%M:%S')
                        break_end_time = datetime.datetime.strptime(break_end, '%Y-%m-%d %H:%M:%S')
                        break_time = (break_end_time - break_start_time).total_seconds() / 3600
                        worked_time -= break_time

                    total_hours_with_breaks += worked_time

            # Write to weekly summary if the week number matches
            if week_number == current_week:
                summary_sheet.append([current_week, total_hours_with_breaks, total_hours_without_breaks])

    wb.save(EXCEL_FILE)

# Tkinter GUI Setup
app = tk.Tk()
app.title("Attendance System")

# Initialize Excel files
init_employee_data_excel()  # Ensure employee data file is created first
employee_names = load_employee_names()  # Now load employee names
init_attendance_excel()  # Then initialize the attendance file

# Listbox for employee selection
employee_listbox = Listbox(app, height=10, width=50, font=("Helvetica", 14))  # Increase font size
employee_listbox.pack(pady=10)

# Add a scrollbar to the listbox
scrollbar = Scrollbar(app)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
employee_listbox.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=employee_listbox.yview)

# Populate the listbox with employee names
for name in employee_names:
    employee_listbox.insert(tk.END, name)

# Main menu buttons
tk.Button(app, text="Clock In", command=lambda: update_employee_action("Clock In")).pack(pady=10)
tk.Button(app, text="Break Start", command=lambda: update_employee_action("Break Start")).pack(pady=10)
tk.Button(app, text="Break End", command=lambda: update_employee_action("Break End")).pack(pady=10)
tk.Button(app, text="Shift End", command=lambda: update_employee_action("Shift End")).pack(pady=10)

# New Employee Registration
tk.Label(app, text="Full Name:").pack(pady=5)
full_name_entry = tk.Entry(app)
full_name_entry.pack(pady=5)

def register_new_employee(event=None):
    full_name = full_name_entry.get().strip()
    
    if not full_name:
        messagebox.showerror("Error", "Please enter a name.")
        return

    name_parts = full_name.split(' ', 1)  # Split the full name into parts
    first_name = name_parts[0]  # Always take the first part as first name
    last_name = name_parts[1] if len(name_parts) > 1 else ""  # If there's a second part, take it as last name, else empty

    # Save to employee data Excel
    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active
    ws.append([first_name, last_name, "", ""])  # ImagePath and FaceEncoding are left empty
    wb.save(EMPLOYEE_DATA_FILE)

    # Update the listbox with the new employee name
    employee_listbox.insert(tk.END, f"{first_name} {last_name}".strip())  # Strip any extra spaces
    messagebox.showinfo("Success", f"Employee {first_name} {last_name.strip()} registered")
    full_name_entry.delete(0, tk.END)

# Bind the Enter key to register a new employee
full_name_entry.bind("<Return>", register_new_employee)
tk.Button(app, text="Register New Employee", command=register_new_employee).pack(pady=20)


# Function to open schedule input page
def open_schedule_input():
    schedule_window = tk.Toplevel(app)
    schedule_window.title("Manual Schedule Input")

    tk.Label(schedule_window, text="Select Employee for Schedule", font=("Helvetica", 14)).pack(pady=10)

    # Listbox for employee selection
    schedule_employee_listbox = Listbox(schedule_window, height=10, width=50, font=("Helvetica", 14))  # Increase font size
    schedule_employee_listbox.pack(pady=10)

    # Add a scrollbar to the listbox
    schedule_scrollbar = Scrollbar(schedule_window)
    schedule_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    schedule_employee_listbox.config(yscrollcommand=schedule_scrollbar.set)
    schedule_scrollbar.config(command=schedule_employee_listbox.yview)

    # Populate the schedule listbox with employee names
    for name in employee_names:
        schedule_employee_listbox.insert(tk.END, name)

    tk.Label(schedule_window, text="Select Your Schedule", font=("Helvetica", 14)).pack(pady=10)

    # Start Time Input
    tk.Label(schedule_window, text="Start Hour").pack()
    start_hour_slider = tk.Scale(schedule_window, from_=1, to=12, orient=tk.HORIZONTAL)
    start_hour_slider.pack()

    tk.Label(schedule_window, text="Start Minute").pack()
    start_minute_slider = tk.Scale(schedule_window, from_=0, to=45, resolution=15, orient=tk.HORIZONTAL)  # 15-minute increments
    start_minute_slider.pack()

    # AM/PM selection for start time
    start_period_var = StringVar()
    frame1 = tk.Frame(schedule_window)
    frame1.pack()
    tk.Radiobutton(frame1, text="AM", variable=start_period_var, value="AM").pack(side=tk.LEFT)
    tk.Radiobutton(frame1, text="PM", variable=start_period_var, value="PM").pack(side=tk.LEFT)

    # End Time Input
    tk.Label(schedule_window, text="End Hour").pack()
    end_hour_slider = tk.Scale(schedule_window, from_=1, to=12, orient=tk.HORIZONTAL)
    end_hour_slider.pack()

    tk.Label(schedule_window, text="End Minute").pack()
    end_minute_slider = tk.Scale(schedule_window, from_=0, to=45, resolution=15, orient=tk.HORIZONTAL)  # 15-minute increments
    end_minute_slider.pack()

    # AM/PM selection for end time
    end_period_var = StringVar()
    frame2 = tk.Frame(schedule_window)
    frame2.pack()
    tk.Radiobutton(frame2, text="AM", variable=end_period_var, value="AM").pack(side=tk.LEFT)
    tk.Radiobutton(frame2, text="PM", variable=end_period_var, value="PM").pack(side=tk.LEFT)

    # Checkbox for break
    break_var = tk.BooleanVar()
    tk.Checkbutton(schedule_window, text="WILL TAKE 30 MIN BREAK", variable=break_var).pack()

    def submit_schedule():
        selected_employee = schedule_employee_listbox.get(tk.ACTIVE)

        if not selected_employee:
            messagebox.showerror("Error", "Please select an employee for the schedule.")
            return

        start_hour = start_hour_slider.get()
        start_minute = start_minute_slider.get()
        end_hour = end_hour_slider.get()
        end_minute = end_minute_slider.get()

        start_period = start_period_var.get()
        end_period = end_period_var.get()

        # Check if AM/PM is selected for both times
        if not start_period or not end_period:
            messagebox.showerror("Error", "Please select AM or PM for both start and end times.")
            return

        # Convert 12-hour format to 24-hour format for comparison
        start_time = (start_hour % 12) + (12 if start_period == "PM" else 0), start_minute
        end_time = (end_hour % 12) + (12 if end_period == "PM" else 0), end_minute

        # Convert to minutes since midnight for easy comparison
        start_time_minutes = start_time[0] * 60 + start_time[1]
        end_time_minutes = end_time[0] * 60 + end_time[1]

        # Prevent invalid shifts
        if end_time_minutes <= start_time_minutes:
            messagebox.showerror("Error", "End time must be after start time.")
            return

        # Calculate total shift time
        total_shift_time = end_time_minutes - start_time_minutes

        # Deduct break time if checked
        if break_var.get():
            if total_shift_time > 120:  # Only apply break if total shift is more than 2 hours
                total_shift_time -= 30  # Deduct 30 minutes for the break

                # Calculate break times
                break_start_minutes = start_time_minutes + 120  # Break starts after 2 hours
                break_end_minutes = break_start_minutes + 30  # Break lasts for 30 minutes

                break_start_hour = break_start_minutes // 60
                break_start_minute = break_start_minutes % 60
                break_end_hour = break_end_minutes // 60
                break_end_minute = break_end_minutes % 60

                # Format break times for saving
                break_start_time = f"{datetime.datetime.now().strftime('%Y-%m-%d')} {break_start_hour % 24}:{break_start_minute:02d}"
                break_end_time = f"{datetime.datetime.now().strftime('%Y-%m-%d')} {break_end_hour % 24}:{break_end_minute:02d}"
            else:
                break_start_time = None
                break_end_time = None
        else:
            break_start_time = None
            break_end_time = None

        # Custom confirmation popup
        confirmation_message = f"CLOCK IN FOR: {selected_employee}\n" \
                               f"AT TIMES: {start_hour}:{start_minute:02d} {start_period} - " \
                               f"{end_hour}:{end_minute:02d} {end_period}?"
        
        if messagebox.askyesno("Confirm Schedule", confirmation_message):
            # Save the schedule in Excel
            wb = load_workbook(EXCEL_FILE)
            current_date = datetime.datetime.now().strftime('%Y-%m-%d')
            ws = wb[current_date]

            # Check if employee already has an entry, if not create one
            employee_row = None
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
                if row[0].value == selected_employee:
                    employee_row = row
                    break
            
            if not employee_row:
                employee_row = [selected_employee, None, None, None, None]  # Initialize row
                ws.append(employee_row)
                row_index = ws.max_row  # The index of the newly appended row
            else:
                row_index = employee_row[0].row  # Get the existing row index
            
            # Update the shift start time
            ws.cell(row=row_index, column=2).value = f"{datetime.datetime.now().strftime('%Y-%m-%d')} {start_hour}:{start_minute:02d} {start_period}"
            # Update the shift end time
            ws.cell(row=row_index, column=5).value = f"{datetime.datetime.now().strftime('%Y-%m-%d')} {end_hour}:{end_minute:02d} {end_period}"

            # If break times are calculated, save them too
            if break_start_time and break_end_time:
                ws.cell(row=row_index, column=3).value = break_start_time  # Break start time
                ws.cell(row_index, column=4).value = break_end_time  # Break end time

            wb.save(EXCEL_FILE)
            messagebox.showinfo("Success", f"Total Shift Time for {selected_employee}: {total_shift_time} minutes")

            # Update summaries after submitting the schedule
            update_summaries(current_date)

    # Add the Submit button after defining the submit_schedule function
    tk.Button(schedule_window, text="Submit Schedule", command=submit_schedule).pack(pady=20)

    # Main Screen button
    tk.Button(schedule_window, text="Main Screen", command=schedule_window.destroy).pack(pady=10)

    schedule_window.transient(app)
    schedule_window.grab_set()
    app.wait_window(schedule_window)

# Function to update the summaries after each schedule submission
def update_summaries(current_date):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[current_date]

    # Initialize summary variables
    total_time = 0
    total_employees = 0

    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[1] is not None and row[4] is not None:  # Check if both start and end times exist
            # Calculate total shift time
            start_time = datetime.datetime.strptime(row[1], '%Y-%m-%d %I:%M %p')
            end_time = datetime.datetime.strptime(row[4], '%Y-%m-%d %I:%M %p')
            shift_time = (end_time - start_time).total_seconds() / 60  # Shift time in minutes
            total_time += shift_time
            total_employees += 1

    # Print summaries
    print(f"Total Employees: {total_employees}")
    print(f"Total Shift Time: {total_time} minutes")

# Button to open schedule input page
tk.Button(app, text="Manual Schedule Input", command=open_schedule_input).pack(pady=20)

app.mainloop()
