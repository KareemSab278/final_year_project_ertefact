import cv2
import face_recognition
import os
import sys
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import datetime
import numpy as np
import time

# Global constants
EXCEL_FILE = 'attendance.xlsx'  # For daily attendance
EMPLOYEE_DATA_FILE = 'employee_data.xlsx'  # For face encodings and image paths
IMAGE_DIR = 'employee_images/'

if not os.path.exists(IMAGE_DIR):
    os.makedirs(IMAGE_DIR)

# Set the path to the face recognition models
def set_face_recognition_model_path():
    model_path = "C:\\Users\\user\\Desktop\\BSc\\final year project\\artefact\\face_recognition_models"
    if os.path.exists(model_path):
        os.environ["FACE_RECOGNITION_MODEL_PATH"] = model_path
    else:
        messagebox.showerror("Error", "Face recognition model path does not exist.")

# Initialise Excel with a new sheet for each day
def init_attendance_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)
    
    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
        wb.save(EXCEL_FILE)

# Initialise employee data Excel (face encodings and image paths)
def init_employee_data_excel():
    if not os.path.exists(EMPLOYEE_DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "EmployeeData"
        ws.append(["First Name", "Last Name", "ImagePath", "FaceEncoding"])
        wb.save(EMPLOYEE_DATA_FILE)

# Optimise camera handling for recognition
def open_camera_for_recognition(action_type):
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        messagebox.showerror("Error", "Cannot access camera. Please check your device.")
        return

    # Set lower resolution for faster performance
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    # Load known face encodings and names
    known_face_encodings, known_face_names = load_registered_faces()

    if not known_face_encodings:
        messagebox.showerror("Error", "No registered faces found. Please register employees first.")
        cap.release()
        return

    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    face_found = False
    process_frame = True  # Skip every alternate frame

    while True:
        ret, frame = cap.read()

        if not ret:
            messagebox.showerror("Error", "Failed to capture image from camera.")
            break

        if process_frame:
            process_frame = False

            # Convert to grayscale for faster Haar Cascade detection
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            if len(faces) > 0:
                # Convert to RGB for face recognition
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                face_locations = face_recognition.face_locations(rgb_frame, model="hog")
                face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

                for face_encoding, face_location in zip(face_encodings, face_locations):
                    matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.6)
                    face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)

                    if True in matches:
                        best_match_index = np.argmin(face_distances)
                        employee_name = known_face_names[best_match_index]
                        face_found = True
                        update_employee_action(employee_name, action_type)
                        messagebox.showinfo("Success", f"{action_type} recorded for {employee_name}")
                        break

        else:
            process_frame = True

        if face_found or cv2.waitKey(1) & 0xFF == ord('q'):
            break

        cv2.imshow('Live Recognition', frame)

    cap.release()
    cv2.destroyAllWindows()

    if not face_found:
        messagebox.showerror("No Match", "No face match found!")

# Load registered faces for recognition
def load_registered_faces():
    known_face_encodings = []
    known_face_names = []

    wb = load_workbook(EMPLOYEE_DATA_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name, last_name, _, face_encoding = row
        full_name = f"{first_name} {last_name}"

        if face_encoding:
            encoding = np.fromstring(face_encoding[1:-1], dtype=float, sep=',')
            known_face_encodings.append(encoding)
            known_face_names.append(full_name)

    return known_face_encodings, known_face_names

# Register a new employee
def register_new_employee():
    first_name = first_name_entry.get()
    last_name = last_name_entry.get()

    if not first_name or not last_name:
        messagebox.showerror("Error", "Please enter both first and last names")
        return

    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)

        if face_locations:
            face_encoding = face_recognition.face_encodings(rgb_frame, face_locations)[0]
            image_filename = f"{first_name}_{last_name}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            image_path = os.path.join(IMAGE_DIR, image_filename)
            cv2.imwrite(image_path, frame)

            # Save to employee data Excel
            wb = load_workbook(EMPLOYEE_DATA_FILE)
            ws = wb.active
            ws.append([first_name, last_name, image_path, str(face_encoding.tolist())])
            wb.save(EMPLOYEE_DATA_FILE)

            messagebox.showinfo("Success", f"Face registered for {first_name} {last_name}")
            first_name_entry.delete(0, tk.END)
            last_name_entry.delete(0, tk.END)
            break

    cap.release()
    cv2.destroyAllWindows()

from openpyxl import Workbook, load_workbook

# Update employee action (clock-in, break start, etc.)
def update_employee_action(full_name, action_type):
    wb = load_workbook(EXCEL_FILE)
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # Check if today's date already has a sheet; if not, create it
    if current_date not in wb.sheetnames:
        ws = wb.create_sheet(current_date)
        ws.append(["Employee Name", "Shift Start", "Break Start", "Break End", "Shift End"])
    else:
        ws = wb[current_date]

    # Check if the employee already has an entry in today's sheet
    employee_row = None
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value == full_name:
            employee_row = row
            break

    # If employee doesn't have an entry, add a new row
    if not employee_row:
        ws.append([full_name, None, None, None, None])
        row_index = ws.max_row
    else:
        row_index = employee_row[0].row

    # Update the appropriate column based on the action type
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if action_type == "Clock In":
        if ws.cell(row=row_index, column=2).value is None:
            ws.cell(row=row_index, column=2).value = timestamp
        else:
            messagebox.showinfo("Already Clocked In", f"{full_name} has already clocked in today.")
            return
    elif action_type == "Break Start":
        if ws.cell(row=row_index, column=3).value is None:
            ws.cell(row=row_index, column=3).value = timestamp
        else:
            messagebox.showinfo("Already Started Break", f"{full_name} has already started break today.")
            return
    elif action_type == "Break End":
        if ws.cell(row=row_index, column=4).value is None:
            ws.cell(row=row_index, column=4).value = timestamp
        else:
            messagebox.showinfo("Already Ended Break", f"{full_name} has already ended break today.")
            return
    elif action_type == "Shift End":
        if ws.cell(row=row_index, column=5).value is None:
            ws.cell(row=row_index, column=5).value = timestamp
        else:
            messagebox.showinfo("Already Clocked Out", f"{full_name} has already ended their shift today.")
            return
    # Save the workbook
    wb.save(EXCEL_FILE)

# Tkinter GUI Setup
app = tk.Tk()
app.title("Employee Attendance System")

# Set the path to the face recognition models before starting the app
set_face_recognition_model_path()

# Main menu buttons
tk.Button(app, text="Clock In", command=lambda: open_camera_for_recognition("Clock In")).pack(pady=10)
tk.Button(app, text="Break Start", command=lambda: open_camera_for_recognition("Break Start")).pack(pady=10)
tk.Button(app, text="Break End", command=lambda: open_camera_for_recognition("Break End")).pack(pady=10)
tk.Button(app, text="Shift End", command=lambda: open_camera_for_recognition("Shift End")).pack(pady=10)

# New Employee Registration
tk.Label(app, text="First Name:").pack(pady=5)
first_name_entry = tk.Entry(app)
first_name_entry.pack(pady=5)

tk.Label(app, text="Last Name:").pack(pady=5)
last_name_entry = tk.Entry(app)
last_name_entry.pack(pady=5)

tk.Button(app, text="Register New Employee", command=register_new_employee).pack(pady=20)

# Initialize Excel files
init_attendance_excel()
init_employee_data_excel()

app.mainloop()
